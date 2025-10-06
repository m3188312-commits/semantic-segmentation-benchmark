import os
import sys
import argparse
import numpy as np
from glob import glob
from PIL import Image
from sklearn.metrics import precision_recall_fscore_support
from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname(__file__)
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..'))
sys.path.insert(0, PROJECT_ROOT)

from models.random_forest.dataset import (
    load_dataset,
    rgb_to_mask,
    COLOR2CLASS,
    extract_features
)
from models.random_forest.model import load_model

CLASS_NAMES = [
    'Unknown', 'Artificial', 'Woodland', 'Arable',
    'Frygana', 'Bareland', 'Water', 'Permanent'
]


def predict_and_save(clf, img_dir: str, out_dir: str, single_image: str = None):
    os.makedirs(out_dir, exist_ok=True)
    inv_map = {v: k for k, v in COLOR2CLASS.items()}

    if single_image:
        img_paths = [os.path.join(img_dir, single_image)]
    else:
        img_paths = sorted(
            glob(os.path.join(img_dir, '*.png')) +
            glob(os.path.join(img_dir, '*.jpg')) +
            glob(os.path.join(img_dir, '*.jpeg'))
        )
    if not img_paths:
        raise FileNotFoundError(f"No images found in {img_dir}")

    for img_path in img_paths:
        img = np.array(Image.open(img_path))
        feats = extract_features(img)
        h, w, f = feats.shape
        preds = clf.predict(feats.reshape(-1, f)).reshape(h, w)

        rgb = np.zeros((h, w, 3), dtype=np.uint8)
        for cls_idx, color in inv_map.items():
            rgb[preds == cls_idx] = color

        base = os.path.basename(img_path).rsplit('.', 1)[0] + '.png'
        Image.fromarray(rgb).save(os.path.join(out_dir, base))


def evaluate_split(pred_dir: str, gt_img_dir: str, gt_mask_dir: str, single_image: str = None):
    if single_image is None:
        _, true_masks = load_dataset(gt_img_dir, gt_mask_dir)
        pred_paths = sorted(glob(os.path.join(pred_dir, '*.png')))
        if len(pred_paths) != len(true_masks):
            raise RuntimeError(
                f"Count mismatch: {len(pred_paths)} preds vs {len(true_masks)} GT masks"
            )
        y_true = []
        y_pred = []
        for gt, pf in zip(true_masks, pred_paths):
            y_true.append(gt.flatten())
            pm = rgb_to_mask(np.array(Image.open(pf).convert('RGB')))
            y_pred.append(pm.flatten())
        y_true = np.hstack(y_true)
        y_pred = np.hstack(y_pred)
    else:
        base = os.path.splitext(single_image)[0]
        gt_pattern = os.path.join(gt_mask_dir, base + '.*')
        gt_files = glob(gt_pattern)
        if not gt_files:
            raise FileNotFoundError(f"No GT mask found for {single_image} in {gt_mask_dir}")
        gt_path = gt_files[0]
        gt_mask = rgb_to_mask(np.array(Image.open(gt_path).convert('RGB')))
        y_true = gt_mask.flatten()
        pred_path = os.path.join(pred_dir, base + '.png')
        if not os.path.isfile(pred_path):
            raise FileNotFoundError(f"No prediction found for {single_image} in {pred_dir}")
        pred_mask = rgb_to_mask(np.array(Image.open(pred_path).convert('RGB')))
        y_pred = pred_mask.flatten()

    p, r, f1, _ = precision_recall_fscore_support(
        y_true, y_pred,
        labels=list(range(len(CLASS_NAMES))),
        zero_division=0
    )
    return {'p': p, 'r': r, 'f1': f1}


def main():
    parser = argparse.ArgumentParser(description="RF evaluation: predict+evaluate with terminal output.")
    parser.add_argument("--evaluate", action="store_true", help="Skip prediction; only compute metrics.")
    parser.add_argument("--predict", action="store_true", help="Only generate segmentations.")
    parser.add_argument("--single-image", nargs=2, metavar=('SPLIT', 'IMAGE'),
                        help="Evaluate a single IMAGE from SPLIT (train, lowres, test).")
    parser.add_argument("--model_path", type=str, help="path to model weights file")
    args = parser.parse_args()
    
    do_pred = args.predict or not args.evaluate
    do_eval = args.evaluate or not args.predict

    data_root = os.path.join(PROJECT_ROOT, 'dataset')
    scripts_dir = os.path.join(PROJECT_ROOT, 'scripts')
    preds_root = os.path.join(PROJECT_ROOT, 'predictions', 'random_forest')

    if args.single_image:
        splits = [args.single_image[0]]
    else:
        splits = ['train', 'test', 'lowres']
    
    results = {}
    for split in splits:
        if args.model_path:
            model_path = args.model_path
        else:
            model_path = os.path.join(scripts_dir, 'rf_model_train.pkl')

        if not os.path.exists(model_path):
            print(f"❌ Model weights not found at: {model_path}")
            if os.path.exists(scripts_dir):
                print("Available files in scripts folder:")
                for f in os.listdir(scripts_dir):
                    if f.endswith('.pkl'):
                        print(f"  - {f}")
            continue
            
        clf = load_model(model_path)
        img_dir = os.path.join(data_root, split, 'image')
        mask_dir = os.path.join(data_root, split, 'mask')
        if not os.path.isdir(mask_dir):
            mask_dir = os.path.join(data_root, split, 'masks')
        pred_dir = os.path.join(preds_root, split)
        
        if do_pred:
            print(f"Generating predictions for '{split}' -> {pred_dir}")
            predict_and_save(clf, img_dir, pred_dir)
        if do_eval:
            print(f"Computing metrics for '{split}'")
            res = evaluate_split(pred_dir, img_dir, mask_dir)
            results[split] = res

            print(f"\n📊 Metrics for '{split}':")
            print(f"{'Class':<12}{'P':>6}{'R':>6}{'F1':>6}")
            print('-'*30)
            for idx, name in enumerate(CLASS_NAMES):
                print(f"{name:<12}{res['p'][idx]:6.3f}{res['r'][idx]:6.3f}{res['f1'][idx]:6.3f}")
    
    if len(results) > 1:
        print(f"\n📋 SUMMARY:")
        for split, res in results.items():
            print(f"{split}: P={res['p'].mean():.3f}, R={res['r'].mean():.3f}, F1={res['f1'].mean():.3f}")

    if results:
        if args.model_path:
            model_name = os.path.basename(args.model_path)
        else:
            model_name = "rf_model_train.pkl"
        save_summary_to_excel(results, model_name)


def save_summary_to_excel(results, model_name):
    out_file = "scripts/rf_results_detailed.xlsx"

    # Extract percent from filename if possible
    percent = "N/A"
    if "pct" in model_name:
        try:
            percent = model_name.split("train_")[1].split("pct")[0] + "%"
        except:
            percent = "N/A"

    # If file exists, open it. Otherwise, create new.
    if os.path.exists(out_file):
        wb = load_workbook(out_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Percent", "Split", "Class", "Precision", "Recall", "F1"])

    for split, res in results.items():
        for idx, class_name in enumerate(CLASS_NAMES):
            ws.append([
                percent,
                split.capitalize(),
                class_name,
                float(res['p'][idx]),
                float(res['r'][idx]),
                float(res['f1'][idx])
            ])
        ws.append([
            percent,
            split.capitalize(),
            "SUMMARY",
            float(res['p'].mean()),
            float(res['r'].mean()),
            float(res['f1'].mean())
        ])
        ws.append([])  # empty row

    wb.save(out_file)
    print(f"[Excel] Results saved → {out_file}")


if __name__ == '__main__':
    main()
